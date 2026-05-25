#!/usr/bin/env python3
from __future__ import annotations

from itertools import product
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path("reports/RealLsu_signal_widths_BSD_CONFIG1.xlsx")


def clog2(n: int) -> int:
    res = 0
    while n > (1 << res):
        res += 1
    return res


def bit_width_for_count(count: int) -> int:
    width = 0
    max_value = count - 1 if count > 0 else 0
    while True:
        width += 1
        max_value >>= 1
        if max_value == 0:
            return width


FETCH_WIDTH = 16
DECODE_WIDTH = 8
COMMIT_WIDTH = DECODE_WIDTH
ARF_NUM = 32
PRF_NUM = 512
MAX_BR_NUM = 128
ROB_NUM = 512
STQ_SIZE = 512
LDQ_SIZE = 512
FTQ_SIZE = 512

LSU_STA_COUNT = 4
LSU_LDU_COUNT = 4
LSU_AGU_COUNT = LSU_STA_COUNT + LSU_LDU_COUNT
LSU_LDU_WIDTH = clog2(LSU_LDU_COUNT)
LSU_SDU_COUNT = 4
LSU_LOAD_WB_WIDTH = LSU_LDU_COUNT

MAX_STQ_DISPATCH_WIDTH = DECODE_WIDTH
MAX_LDQ_DISPATCH_WIDTH = DECODE_WIDTH

AREG_IDX_WIDTH = 6
PRF_IDX_WIDTH = clog2(PRF_NUM)
ROB_IDX_WIDTH = clog2(ROB_NUM)
STQ_IDX_WIDTH = clog2(STQ_SIZE)
LDQ_IDX_WIDTH = clog2(LDQ_SIZE)
BR_TAG_WIDTH = clog2(MAX_BR_NUM)
BR_MASK_WIDTH = MAX_BR_NUM
CSR_IDX_WIDTH = 12
FTQ_IDX_WIDTH = clog2(FTQ_SIZE)
FTQ_OFFSET_WIDTH = clog2(FETCH_WIDTH)
INST_TYPE_WIDTH = bit_width_for_count(20)
UOP_TYPE_WIDTH = bit_width_for_count(18)
ROB_CPLT_MASK_WIDTH = 3
LDQ_STQ_IDX_WIDTH = clog2(LDQ_SIZE + STQ_SIZE)
MAX_IDX_WIDTH = clog2(max(LDQ_SIZE, STQ_SIZE))
LSU_MMU_IDX_WIDTH = max(LDQ_IDX_WIDTH, STQ_IDX_WIDTH)


def dim(name: str, count: int) -> tuple[str, int]:
    return name, count


def count_from_dims(dims: list[tuple[str, int]]) -> int:
    total = 1
    for _, count in dims:
        total *= count
    return total


def compact_path(path: str, dims: list[tuple[str, int]]) -> str:
    out = path
    for name, count in dims:
        out = out.replace("{" + name + "}", f"0:{count - 1}")
    return out


def expand_path(path: str, dims: list[tuple[str, int]]) -> list[str]:
    if not dims:
        return [path]
    names = [name for name, _ in dims]
    ranges = [range(count) for _, count in dims]
    expanded = []
    for values in product(*ranges):
        cur = path
        for name, value in zip(names, values):
            cur = cur.replace("{" + name + "}", str(value))
        expanded.append(cur)
    return expanded


def add(rows, scope, path, type_name, width, dims=None, note=""):
    dims = dims or []
    rows.append(
        {
            "Scope": scope,
            "Signal": compact_path(path, dims),
            "Type": type_name,
            "Width(bits)": width,
            "Array count": count_from_dims(dims),
            "Total bits": width * count_from_dims(dims),
            "Note": note,
            "_path": path,
            "_dims": dims,
        }
    )


def add_debug_meta(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.instruction", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.pc", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.mem_align_mask", "uint8_t", 8, dims)
    add(rows, scope, f"{prefix}.difftest_skip", "bool", 1, dims)
    add(rows, scope, f"{prefix}.inst_idx", "int64_t", 64, dims)


def add_tma_meta(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.is_cache_miss", "bool", 1, dims)
    add(rows, scope, f"{prefix}.is_ret", "bool", 1, dims)
    add(rows, scope, f"{prefix}.mem_commit_is_load", "bool", 1, dims)
    add(rows, scope, f"{prefix}.mem_commit_is_store", "bool", 1, dims)


def add_csr_status(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.sstatus", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.mstatus", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.satp", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.privilege", "wire<2>", 2, dims)


def add_mmu_resp(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.paddr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.result", "MMUResultType(wire<2>)", 2, dims)
    add(rows, scope, f"{prefix}.entry_idx", "wire<LSU_MMU_IDX_WIDTH>", LSU_MMU_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.wait_idx", "wire<LSU_MMU_IDX_WIDTH>", LSU_MMU_IDX_WIDTH, dims)


def add_mmu_req(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.vaddr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.entry_idx", "wire<LSU_MMU_IDX_WIDTH>", LSU_MMU_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.wait_idx", "wire<LSU_MMU_IDX_WIDTH>", LSU_MMU_IDX_WIDTH, dims)


def add_load_resp(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.data", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.req_id", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.replay", "ReplayType(wire<2>)", 2, dims)


def add_store_resp(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.replay", "ReplayType(wire<2>)", 2, dims)
    add(rows, scope, f"{prefix}.req_id", "wire<32>", 32, dims)


def add_load_req(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.addr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.req_id", "wire<32>", 32, dims)


def add_store_req(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.addr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.data", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.strb", "wire<8>", 8, dims)
    add(rows, scope, f"{prefix}.req_id", "wire<32>", 32, dims)


def add_rob_commit_inst(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.diag_val", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.dest_areg", "wire<AREG_IDX_WIDTH>", AREG_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.dest_preg", "wire<PRF_IDX_WIDTH>", PRF_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.old_dest_preg", "wire<PRF_IDX_WIDTH>", PRF_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.ftq_idx", "wire<FTQ_IDX_WIDTH>", FTQ_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.ftq_offset", "wire<FTQ_OFFSET_WIDTH>", FTQ_OFFSET_WIDTH, dims)
    add(rows, scope, f"{prefix}.ftq_is_last", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.mispred", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.br_taken", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.dest_en", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.func7", "wire<7>", 7, dims)
    add(rows, scope, f"{prefix}.rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_flag", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.stq_idx", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.stq_flag", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.page_fault_inst", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.page_fault_load", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.page_fault_store", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.illegal_inst", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.type", "wire<INST_TYPE_WIDTH>", INST_TYPE_WIDTH, dims)
    add_tma_meta(rows, scope, f"{prefix}.tma", dims)
    add_debug_meta(rows, scope, f"{prefix}.dbg", dims)
    add(rows, scope, f"{prefix}.flush_pipe", "wire<1>", 1, dims)


def add_exe_lsu_req_uop(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.result", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.dest_preg", "wire<PRF_IDX_WIDTH>", PRF_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.func3", "wire<3>", 3, dims)
    add(rows, scope, f"{prefix}.func7", "wire<7>", 7, dims)
    add(rows, scope, f"{prefix}.is_atomic", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.br_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.stq_idx", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.stq_flag", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.ldq_idx", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_flag", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.dest_en", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.op", "wire<UOP_TYPE_WIDTH>", UOP_TYPE_WIDTH, dims)
    add_debug_meta(rows, scope, f"{prefix}.dbg", dims)


def add_lsu_exe_resp_uop(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.diag_val", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.result", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.dest_preg", "wire<PRF_IDX_WIDTH>", PRF_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.br_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.dest_en", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.page_fault_load", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.page_fault_store", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.op", "wire<UOP_TYPE_WIDTH>", UOP_TYPE_WIDTH, dims)
    add_debug_meta(rows, scope, f"{prefix}.dbg", dims)
    add(rows, scope, f"{prefix}.flush_pipe", "wire<1>", 1, dims)


def add_ldq_entry(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.load_state", "LoadState(uint8_t)", 8, dims)
    add(rows, scope, f"{prefix}.v_addr_valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.v_addr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.p_addr_valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.p_addr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.replay_type", "ReplayType(wire<2>)", 2, dims)
    add(rows, scope, f"{prefix}.func3", "wire<3>", 3, dims)
    add(rows, scope, f"{prefix}.result", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.dest_preg", "wire<PRF_IDX_WIDTH>", PRF_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.byte_mask", "wire<4>", 4, dims)
    add(rows, scope, f"{prefix}.rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_flag", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.br_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH, dims)
    add(rows, scope, f"{prefix}.is_mmio", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.diag_val", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.page_fault", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.is_lrsc", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.stq_snapshot.idx", "reg<STQ_IDX_WIDTH>", STQ_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.stq_snapshot.flag", "reg<1>", 1, dims)


def add_stq_entry(rows, scope, prefix, dims=None):
    dims = dims or []
    add(rows, scope, f"{prefix}.data_valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.data", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.func3", "wire<3>", 3, dims)
    add(rows, scope, f"{prefix}.vaddr_valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.vaddr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.paddr_valid", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.paddr", "wire<32>", 32, dims)
    add(rows, scope, f"{prefix}.page_fault", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.suppress_write", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.is_mmio", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.is_lrsc", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.sc_pass", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.dest_preg", "wire<PRF_IDX_WIDTH>", PRF_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.store_state", "StoreState(uint8_t)", 8, dims)
    add(rows, scope, f"{prefix}.replay_type", "ReplayType(wire<2>)", 2, dims)
    add(rows, scope, f"{prefix}.replay_wait_cycles", "uint32_t", 32, dims)
    add(rows, scope, f"{prefix}.br_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, dims)
    add(rows, scope, f"{prefix}.rob_flag", "wire<1>", 1, dims)
    add(rows, scope, f"{prefix}.stq_flag", "wire<1>", 1, dims)


def add_state(rows, scope):
    add_ldq_entry(rows, scope, f"{scope}.ldq[{{i}}]", [dim("i", LDQ_SIZE)])
    add_stq_entry(rows, scope, f"{scope}.stq[{{i}}]", [dim("i", STQ_SIZE)])
    add(rows, scope, f"{scope}.stq_head", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.stq_commit", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.stq_head_flag", "wire<1>", 1)
    add(rows, scope, f"{scope}.stq_commit_count", "wire<STQ_IDX_WIDTH+1>", STQ_IDX_WIDTH + 1)
    add(rows, scope, f"{scope}.stq_count", "wire<STQ_IDX_WIDTH+1>", STQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.wait_mmu_stq[{{i}}].valid", "wire<1>", 1, [dim("i", STQ_SIZE)])
    add(rows, scope, f"{scope}.wait_mmu_stq[{{i}}].stq_idx", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH, [dim("i", STQ_SIZE)])
    add(rows, scope, f"{scope}.wait_mmu_stq_head", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.wait_mmu_stq_count", "wire<STQ_IDX_WIDTH+1>", STQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.wait_mmu_ldq[{{i}}].valid", "wire<1>", 1, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.wait_mmu_ldq[{{i}}].ldq_idx", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.wait_mmu_ldq_head", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.wait_mmu_ldq_count", "wire<LDQ_IDX_WIDTH+1>", LDQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.mmu_done_stq[{{i}}].valid", "wire<1>", 1, [dim("i", STQ_SIZE)])
    add(rows, scope, f"{scope}.mmu_done_stq[{{i}}].stq_idx", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH, [dim("i", STQ_SIZE)])
    add(rows, scope, f"{scope}.mmu_done_stq_head", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.mmu_done_stq_count", "wire<STQ_IDX_WIDTH+1>", STQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.finish[{{i}}].valid", "wire<1>", 1, [dim("i", LDQ_SIZE + STQ_SIZE)])
    add(rows, scope, f"{scope}.finish[{{i}}].idx", "wire<MAX_IDX_WIDTH>", MAX_IDX_WIDTH, [dim("i", LDQ_SIZE + STQ_SIZE)])
    add(rows, scope, f"{scope}.finish[{{i}}].is_load", "wire<1>", 1, [dim("i", LDQ_SIZE + STQ_SIZE)])
    add(rows, scope, f"{scope}.finish_head", "wire<LDQ_STQ_IDX_WIDTH>", LDQ_STQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.finish_count", "wire<LDQ_STQ_IDX_WIDTH+1>", LDQ_STQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.stlf_queue[{{i}}].valid", "wire<1>", 1, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.stlf_queue[{{i}}].ldq_idx", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.stlf_queue_head", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.stlf_queue_count", "wire<LDQ_IDX_WIDTH+1>", LDQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.wait_dcache_ldq[{{i}}].valid", "wire<1>", 1, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.wait_dcache_ldq[{{i}}].req_gen", "wire<31-LDQ_IDX_WIDTH>", 31 - LDQ_IDX_WIDTH, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.wait_dcache_ldq[{{i}}].ldq_idx", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.wait_dcache_ldq[{{i}}].replay_wait_cycles", "uint32_t", 32, [dim("i", LDQ_SIZE)])
    add(rows, scope, f"{scope}.wait_dcache_ldq_head", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH)
    add(rows, scope, f"{scope}.wait_dcache_ldq_count", "wire<LDQ_IDX_WIDTH+1>", LDQ_IDX_WIDTH + 1)

    add(rows, scope, f"{scope}.uncached_unit.valid", "wire<1>", 1)
    add(rows, scope, f"{scope}.uncached_unit.is_load", "wire<1>", 1)
    add(rows, scope, f"{scope}.uncached_unit.addr", "wire<32>", 32)
    add(rows, scope, f"{scope}.uncached_unit.wdata", "wire<32>", 32)
    add(rows, scope, f"{scope}.uncached_unit.func3", "wire<32>", 32)
    add(rows, scope, f"{scope}.uncached_unit.idx", "wire<MAX_IDX_WIDTH>", MAX_IDX_WIDTH)

    add(rows, scope, f"{scope}.lrsc_unit.reserve_valid", "wire<1>", 1)
    add(rows, scope, f"{scope}.lrsc_unit.reserve_addr", "wire<32>", 32)
    add(rows, scope, f"{scope}.lrsc_unit.reserve_rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH)
    add(rows, scope, f"{scope}.lrsc_unit.reserve_rob_flag", "wire<1>", 1)
    add(rows, scope, f"{scope}.lrsc_unit.reserve_br_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH)

    add(rows, scope, f"{scope}.stq_full_flag", "wire<1>", 1)
    add(rows, scope, f"{scope}.ldq_full_flag", "wire<1>", 1)
    add(rows, scope, f"{scope}.finish_full_flag", "wire<1>", 1)
    add(rows, scope, f"{scope}.req_gen", "wire<31-LDQ_IDX_WIDTH>", 31 - LDQ_IDX_WIDTH)


def build_in_rows():
    rows = []
    scope = "in"
    d = [dim("i", COMMIT_WIDTH)]
    add(rows, scope, "in.rob_commit->commit_entry[{i}].valid", "wire<1>", 1, d)
    add_rob_commit_inst(rows, scope, "in.rob_commit->commit_entry[{i}].uop", d)

    for name in [
        "flush",
        "mret",
        "sret",
        "ecall",
        "exception",
        "fence",
        "fence_i",
        "page_fault_inst",
        "page_fault_load",
        "page_fault_store",
        "illegal_inst",
        "interrupt",
    ]:
        add(rows, scope, f"in.rob_bcast->{name}", "wire<1>", 1)
    add(rows, scope, "in.rob_bcast->trap_val", "wire<32>", 32)
    add(rows, scope, "in.rob_bcast->pc", "wire<32>", 32)
    add(rows, scope, "in.rob_bcast->head_rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH)
    add(rows, scope, "in.rob_bcast->head_valid", "wire<1>", 1)
    add(rows, scope, "in.rob_bcast->head_incomplete_rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH)
    add(rows, scope, "in.rob_bcast->head_incomplete_valid", "wire<1>", 1)

    add(rows, scope, "in.dec_bcast->mispred", "wire<1>", 1)
    add(rows, scope, "in.dec_bcast->br_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH)
    add(rows, scope, "in.dec_bcast->br_id", "wire<BR_TAG_WIDTH>", BR_TAG_WIDTH)
    add(rows, scope, "in.dec_bcast->redirect_rob_idx", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH)
    add(rows, scope, "in.dec_bcast->clear_mask", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH)

    add_csr_status(rows, scope, "in.csr_status->")
    for row in rows:
        row["Signal"] = row["Signal"].replace("->.", "->")
        row["_path"] = row["_path"].replace("->.", "->")

    d = [dim("i", MAX_STQ_DISPATCH_WIDTH)]
    add(rows, scope, "in.dis2lsu->alloc_req[{i}]", "wire<1>", 1, d)
    add(rows, scope, "in.dis2lsu->br_mask[{i}]", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH, d)
    add(rows, scope, "in.dis2lsu->func3[{i}]", "wire<3>", 3, d)
    add(rows, scope, "in.dis2lsu->rob_idx[{i}]", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, d)
    add(rows, scope, "in.dis2lsu->rob_flag[{i}]", "wire<1>", 1, d)
    add(rows, scope, "in.dis2lsu->stq_flag[{i}]", "wire<1>", 1, d)
    d = [dim("i", MAX_LDQ_DISPATCH_WIDTH)]
    add(rows, scope, "in.dis2lsu->ldq_alloc_req[{i}]", "wire<1>", 1, d)
    add(rows, scope, "in.dis2lsu->ldq_idx[{i}]", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH, d)
    add(rows, scope, "in.dis2lsu->ldq_br_mask[{i}]", "wire<BR_MASK_WIDTH>", BR_MASK_WIDTH, d)
    add(rows, scope, "in.dis2lsu->ldq_rob_idx[{i}]", "wire<ROB_IDX_WIDTH>", ROB_IDX_WIDTH, d)
    add(rows, scope, "in.dis2lsu->ldq_rob_flag[{i}]", "wire<1>", 1, d)

    d = [dim("i", LSU_AGU_COUNT)]
    add(rows, scope, "in.exe2lsu->agu_req[{i}].valid", "wire<1>", 1, d)
    add_exe_lsu_req_uop(rows, scope, "in.exe2lsu->agu_req[{i}].uop", d)
    d = [dim("i", LSU_SDU_COUNT)]
    add(rows, scope, "in.exe2lsu->sdu_req[{i}].valid", "wire<1>", 1, d)
    add_exe_lsu_req_uop(rows, scope, "in.exe2lsu->sdu_req[{i}].uop", d)

    add(rows, scope, "in.peripheral_resp->is_mmio", "wire<1>", 1)
    add(rows, scope, "in.peripheral_resp->ready", "wire<1>", 1)
    add(rows, scope, "in.peripheral_resp->mmio_rdata", "wire<32>", 32)

    add_load_resp(rows, scope, "in.dcache2lsu->resp_ports.load_resps[{i}]", [dim("i", LSU_LDU_COUNT)])
    add_store_resp(rows, scope, "in.dcache2lsu->resp_ports.store_resps[{i}]", [dim("i", LSU_STA_COUNT)])
    add(rows, scope, "in.dcache2lsu->mshr_fill", "wire<1>", 1)
    add(rows, scope, "in.dcache2lsu->mshr_fill_addr", "wire<32>", 32)

    add_mmu_resp(rows, scope, "in.mmu2lsu->ldq_resp[{i}]", [dim("i", LSU_LDU_COUNT)])
    add_mmu_resp(rows, scope, "in.mmu2lsu->stq_resp[{i}]", [dim("i", LSU_STA_COUNT)])
    return rows


def build_out_rows():
    rows = []
    scope = "out"
    add(rows, scope, "out.lsu2dis->stq_tail", "wire<STQ_IDX_WIDTH>", STQ_IDX_WIDTH)
    add(rows, scope, "out.lsu2dis->stq_tail_flag", "wire<1>", 1)
    add(rows, scope, "out.lsu2dis->stq_free", "wire<bit_width_for_count(STQ_SIZE+1)>", bit_width_for_count(STQ_SIZE + 1))
    add(rows, scope, "out.lsu2dis->ldq_free", "wire<bit_width_for_count(LDQ_SIZE+1)>", bit_width_for_count(LDQ_SIZE + 1))
    add(rows, scope, "out.lsu2dis->ldq_alloc_idx[{i}]", "wire<LDQ_IDX_WIDTH>", LDQ_IDX_WIDTH, [dim("i", MAX_LDQ_DISPATCH_WIDTH)])
    add(rows, scope, "out.lsu2dis->ldq_alloc_valid[{i}]", "wire<1>", 1, [dim("i", MAX_LDQ_DISPATCH_WIDTH)])

    add(rows, scope, "out.lsu2rob->tma.miss_mask", "std::bitset<ROB_NUM>", ROB_NUM)
    add(rows, scope, "out.lsu2rob->committed_store_pending", "wire<1>", 1)

    add(rows, scope, "out.lsu2exe->wb_req[{i}].valid", "wire<1>", 1, [dim("i", LSU_LOAD_WB_WIDTH)])
    add_lsu_exe_resp_uop(rows, scope, "out.lsu2exe->wb_req[{i}].uop", [dim("i", LSU_LOAD_WB_WIDTH)])
    add(rows, scope, "out.lsu2exe->sta_wb_req[{i}].valid", "wire<1>", 1, [dim("i", LSU_STA_COUNT)])
    add_lsu_exe_resp_uop(rows, scope, "out.lsu2exe->sta_wb_req[{i}].uop", [dim("i", LSU_STA_COUNT)])

    add(rows, scope, "out.peripheral_req->is_mmio", "wire<1>", 1)
    add(rows, scope, "out.peripheral_req->wen", "wire<1>", 1)
    add(rows, scope, "out.peripheral_req->mmio_addr", "wire<32>", 32)
    add(rows, scope, "out.peripheral_req->mmio_wdata", "wire<32>", 32)
    add(rows, scope, "out.peripheral_req->mmio_fun3", "wire<3>", 3)

    add_load_req(rows, scope, "out.lsu2dcache->req_ports.load_ports[{i}]", [dim("i", LSU_LDU_COUNT)])
    add_store_req(rows, scope, "out.lsu2dcache->req_ports.store_ports[{i}]", [dim("i", LSU_STA_COUNT)])
    add(rows, scope, "out.lsu2dcache->icache_req", "wire<LSU_LDU_WIDTH+1>", LSU_LDU_WIDTH + 1)

    add_mmu_req(rows, scope, "out.lsu2mmu->ldq_req[{i}]", [dim("i", LSU_LDU_COUNT)])
    add_mmu_req(rows, scope, "out.lsu2mmu->stq_req[{i}]", [dim("i", LSU_STA_COUNT)])
    add_csr_status(rows, scope, "out.lsu2mmu->csr_status")
    return rows


def clean_csr_arrows(rows):
    for row in rows:
        row["Signal"] = row["Signal"].replace("->.", "->")
        row["_path"] = row["_path"].replace("->.", "->")


def flat_rows(rows):
    out = []
    for row in rows:
        for sig in expand_path(row["_path"], row["_dims"]):
            out.append(
                {
                    "Scope": row["Scope"],
                    "Signal": sig,
                    "Type": row["Type"],
                    "Width(bits)": row["Width(bits)"],
                    "Note": row["Note"],
                }
            )
    return out


def write_sheet(ws, rows, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, 1):
        max_len = len(header)
        for cell in ws[get_column_letter(idx)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def main():
    in_rows = build_in_rows()
    out_rows = build_out_rows()
    cur_rows = []
    nxt_rows = []
    add_state(cur_rows, "cur")
    add_state(nxt_rows, "nxt")

    clean_csr_arrows(in_rows)
    clean_csr_arrows(out_rows)

    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = []
    for name, rows in [
        ("LsuIn in", in_rows),
        ("LsuOut out", out_rows),
        ("LsuState cur", cur_rows),
        ("LsuState nxt", nxt_rows),
    ]:
        summary_rows.append(
            {
                "Object": name,
                "Compact leaf rows": len(rows),
                "Expanded rows": len(flat_rows(rows)),
                "Total bits": sum(r["Total bits"] for r in rows),
                "Note": "BSD_CONFIG=1; pointer fields are expanded to pointed IO payloads.",
            }
        )
    write_sheet(
        wb.create_sheet("Summary"),
        summary_rows,
        ["Object", "Compact leaf rows", "Expanded rows", "Total bits", "Note"],
    )

    const_rows = [
        ("BSD_CONFIG", 1, "Assumed by request"),
        ("DECODE_WIDTH", DECODE_WIDTH, "include/config.h"),
        ("COMMIT_WIDTH", COMMIT_WIDTH, "include/config.h"),
        ("LDQ_SIZE", LDQ_SIZE, "include/config.h"),
        ("STQ_SIZE", STQ_SIZE, "include/config.h"),
        ("LSU_LDU_COUNT", LSU_LDU_COUNT, "count_ports_with_mask(OP_MASK_LD)"),
        ("LSU_STA_COUNT", LSU_STA_COUNT, "count_ports_with_mask(OP_MASK_STA)"),
        ("LSU_AGU_COUNT", LSU_AGU_COUNT, "LSU_LDU_COUNT + LSU_STA_COUNT"),
        ("LSU_SDU_COUNT", LSU_SDU_COUNT, "count_ports_with_mask(OP_MASK_STD)"),
        ("LSU_LOAD_WB_WIDTH", LSU_LOAD_WB_WIDTH, "LSU_LDU_COUNT"),
        ("AREG_IDX_WIDTH", AREG_IDX_WIDTH, "config.h"),
        ("PRF_IDX_WIDTH", PRF_IDX_WIDTH, "clog2(PRF_NUM=512)"),
        ("ROB_IDX_WIDTH", ROB_IDX_WIDTH, "clog2(ROB_NUM=512)"),
        ("LDQ_IDX_WIDTH", LDQ_IDX_WIDTH, "clog2(LDQ_SIZE=512)"),
        ("STQ_IDX_WIDTH", STQ_IDX_WIDTH, "clog2(STQ_SIZE=512)"),
        ("BR_TAG_WIDTH", BR_TAG_WIDTH, "clog2(MAX_BR_NUM=128)"),
        ("BR_MASK_WIDTH", BR_MASK_WIDTH, "MAX_BR_NUM"),
        ("FTQ_IDX_WIDTH", FTQ_IDX_WIDTH, "clog2(FTQ_SIZE=512)"),
        ("FTQ_OFFSET_WIDTH", FTQ_OFFSET_WIDTH, "clog2(FETCH_WIDTH=16)"),
        ("INST_TYPE_WIDTH", INST_TYPE_WIDTH, "bit_width_for_count(20)"),
        ("UOP_TYPE_WIDTH", UOP_TYPE_WIDTH, "bit_width_for_count(MAX_UOP_TYPE=18)"),
        ("ROB_CPLT_MASK_WIDTH", ROB_CPLT_MASK_WIDTH, "base_types.h"),
        ("LDQ_STQ_IDX_WIDTH", LDQ_STQ_IDX_WIDTH, "clog2(LDQ_SIZE+STQ_SIZE)"),
        ("MAX_IDX_WIDTH", MAX_IDX_WIDTH, "clog2(max(LDQ_SIZE,STQ_SIZE))"),
        ("LSU_MMU_IDX_WIDTH", LSU_MMU_IDX_WIDTH, "max(LDQ_IDX_WIDTH,STQ_IDX_WIDTH)"),
        ("req_gen", 31 - LDQ_IDX_WIDTH, "wire<31-LDQ_IDX_WIDTH>"),
    ]
    write_sheet(
        wb.create_sheet("Constants"),
        [{"Name": n, "Value": v, "Source/Expr": s} for n, v, s in const_rows],
        ["Name", "Value", "Source/Expr"],
    )

    compact_headers = ["Scope", "Signal", "Type", "Width(bits)", "Array count", "Total bits", "Note"]
    write_sheet(wb.create_sheet("Compact_In"), in_rows, compact_headers)
    write_sheet(wb.create_sheet("Compact_Out"), out_rows, compact_headers)
    write_sheet(wb.create_sheet("Compact_Cur"), cur_rows, compact_headers)
    write_sheet(wb.create_sheet("Compact_Nxt"), nxt_rows, compact_headers)

    flat_headers = ["Scope", "Signal", "Type", "Width(bits)", "Note"]
    write_sheet(wb.create_sheet("Flat_In"), flat_rows(in_rows), flat_headers)
    write_sheet(wb.create_sheet("Flat_Out"), flat_rows(out_rows), flat_headers)
    write_sheet(wb.create_sheet("Flat_Cur"), flat_rows(cur_rows), flat_headers)
    write_sheet(wb.create_sheet("Flat_Nxt"), flat_rows(nxt_rows), flat_headers)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
