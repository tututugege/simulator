#!/usr/bin/env python3
import argparse
import os
import re
from openpyxl import Workbook


BENCH_RE = re.compile(r"Benchmark:\s+(\S+)(?:\s+\(key=([^)]+)\))?")
NUM_RE = re.compile(r"([-+]?\d+(?:\.\d+)?)")

FIELD_PREFIXES = {
    "cache_hit_rate_pct": "Cache Hit Rate:",
    "l1d_hit_rate_pct": "L1D Hit Rate:",
    "l1i_hit_rate_pct": "L1I Hit Rate:",
    "llc_read_hit_rate_pct": "LLC Read HitRate:",
    "llc_l1i_hit_rate_pct": "LLC L1I HitRate:",
    "llc_l1d_hit_rate_pct": "LLC L1D HitRate:",
    "l1d_avg_miss_penalty_cyc": "L1D Avg Miss Penalty:",
    "l1d_avg_axi_read_cyc": "L1D Avg AXI Read:",
    "l1d_avg_axi_write_cyc": "L1D Avg AXI Write:",
    "l1d_avg_mem_inst_cyc": "L1D Avg Mem-Inst:",
    "l1i_avg_miss_penalty_cyc": "L1I Avg Miss Penalty:",
    "l1i_avg_axi_read_cyc": "L1I Avg AXI Read:",
}

HEADERS = [
    "benchmark",
    "bench_key",
    "cache_hit_rate_pct",
    "l1d_hit_rate_pct",
    "l1i_hit_rate_pct",
    "llc_read_hit_rate_pct",
    "llc_l1i_hit_rate_pct",
    "llc_l1d_hit_rate_pct",
    "l1d_avg_miss_penalty_cyc",
    "l1d_avg_axi_read_cyc",
    "l1d_avg_axi_write_cyc",
    "l1d_avg_mem_inst_cyc",
    "l1i_avg_miss_penalty_cyc",
    "l1i_avg_axi_read_cyc",
]


def _to_float(line):
    tail = line.split(":", 1)[1] if ":" in line else line
    m = NUM_RE.search(tail)
    return float(m.group(1)) if m else None


def parse_report(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    rows = []
    current = None
    for raw in lines:
        line = raw.strip()
        if not line:
            continue

        m = BENCH_RE.search(line)
        if m:
            if current:
                rows.append(current)
            current = {k: None for k in HEADERS}
            current["benchmark"] = m.group(1)
            current["bench_key"] = m.group(2) if m.group(2) else ""
            continue

        if not current:
            continue

        if line.startswith("================================================================="):
            rows.append(current)
            current = None
            continue

        for field, prefix in FIELD_PREFIXES.items():
            if line.startswith(prefix):
                current[field] = _to_float(line)
                break

    if current:
        rows.append(current)

    return rows


def write_xlsx(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "CacheLatency"
    ws.append(HEADERS)

    for row in rows:
        ws.append([row.get(h) for h in HEADERS])

    for col in range(3, len(HEADERS) + 1):
        for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
            for cell in col_cells:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Extract cache hit-rate and latency metrics to Excel.")
    parser.add_argument(
        "-i",
        "--input",
        default="data/2026.3.29.txt",
        help="Input perf report text file.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="data/2026.3.29_cache_latency.xlsx",
        help="Output xlsx path.",
    )
    args = parser.parse_args()

    rows = parse_report(args.input)
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    write_xlsx(rows, args.output)
    print(f"Parsed benchmarks: {len(rows)}")
    print(f"Wrote: {os.path.abspath(args.output)}")


if __name__ == "__main__":
    main()
